from __future__ import annotations

import csv
import json
import math
from datetime import datetime, timedelta, timezone
from io import BytesIO, StringIO
from typing import Any, Final, Literal, Mapping, Sequence

from app.repositories.analytics_repository import (
    AnalyticsRepository,
    analytics_repository,
)
from app.schemas.analytics_schema import (
    AnalyticsCategoryValueSchema,
    AnalyticsChartPointSchema,
    AnalyticsCountCategorySchema,
    AnalyticsCountChartPointSchema,
    AnalyticsDateRangeSchema,
    AnalyticsExportFormat,
    AnalyticsExportRequestSchema,
    AnalyticsExportResponseSchema,
    AnalyticsGranularity,
    AnalyticsOverviewResponseSchema,
    GameAnalyticsResponseSchema,
    GamePerformanceItemSchema,
    RevenueAnalyticsResponseSchema,
    TransactionAnalyticsResponseSchema,
    UserAnalyticsResponseSchema,
    WalletAnalyticsResponseSchema,
    WalletBalanceBucketSchema,
)


AnalyticsSection = Literal[
    "overview",
    "revenue",
    "users",
    "wallet",
    "transactions",
    "games",
    "full",
]

GrowthDirection = Literal["increase", "decrease", "unchanged"]

NumericValue = int | float


UTC: Final = timezone.utc

DEFAULT_ANALYTICS_RANGE_DAYS: Final[int] = 30
DEFAULT_EXPORT_RANGE_DAYS: Final[int] = 30
MAX_ANALYTICS_RANGE_DAYS: Final[int] = 3_650
DEFAULT_GAME_LIMIT: Final[int] = 10
MAX_GAME_LIMIT: Final[int] = 100

SUPPORTED_GRANULARITIES: Final[tuple[AnalyticsGranularity, ...]] = (
    "hour",
    "day",
    "week",
    "month",
)

SUPPORTED_EXPORT_FORMATS: Final[tuple[AnalyticsExportFormat, ...]] = (
    "csv",
    "xlsx",
    "pdf",
)

SUPPORTED_ANALYTICS_SECTIONS: Final[tuple[AnalyticsSection, ...]] = (
    "overview",
    "revenue",
    "users",
    "wallet",
    "transactions",
    "games",
    "full",
)

EXPORT_CONTENT_TYPES: Final[dict[AnalyticsExportFormat, str]] = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    ),
    "pdf": "application/pdf",
}

EXPORT_FILE_EXTENSIONS: Final[dict[AnalyticsExportFormat, str]] = {
    "csv": "csv",
    "xlsx": "xlsx",
    "pdf": "pdf",
}

GRANULARITY_DEFAULT_RANGE_DAYS: Final[
    dict[AnalyticsGranularity, int]
] = {
    "hour": 1,
    "day": 30,
    "week": 180,
    "month": 365,
}

GRANULARITY_MAX_RANGE_DAYS: Final[
    dict[AnalyticsGranularity, int]
] = {
    "hour": 31,
    "day": 3_650,
    "week": 3_650,
    "month": 3_650,
}

ROUNDING_PRECISION: Final[int] = 2

CSV_ENCODING: Final[str] = "utf-8-sig"
CSV_NEWLINE: Final[str] = ""

EXPORT_FILENAME_PREFIX: Final[str] = "analytics"

EMPTY_LABEL: Final[str] = "Unknown"


class AnalyticsServiceError(Exception):
    """Base exception raised by the analytics service."""

    def __init__(
        self,
        message: str,
        *,
        code: str = "analytics_service_error",
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(message)
        self.message = message
        self.code = code
        self.details = dict(details or {})

    def __str__(self) -> str:
        return self.message


class AnalyticsValidationError(AnalyticsServiceError):
    """Raised when analytics input parameters are invalid."""

    def __init__(
        self,
        message: str,
        *,
        code: str = "analytics_validation_error",
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(
            message,
            code=code,
            details=details,
        )


class AnalyticsRepositoryError(AnalyticsServiceError):
    """Raised when analytics data cannot be loaded from the repository."""

    def __init__(
        self,
        message: str = "Unable to retrieve analytics data.",
        *,
        code: str = "analytics_repository_error",
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(
            message,
            code=code,
            details=details,
        )


class AnalyticsExportError(AnalyticsServiceError):

    def __init__(
        self,
        message: str = "Unable to generate the analytics export.",
        *,
        code: str = "analytics_export_error",
        details: Mapping[str, Any] | None = None,
    ) -> None:
        super().__init__(
            message,
            code=code,
            details=details,
        )


class AnalyticsDependencyError(AnalyticsExportError):

    def __init__(
        self,
        dependency: str,
        *,
        export_format: AnalyticsExportFormat,
    ) -> None:
        super().__init__(
            (
                f"The '{dependency}' package is required to generate "
                f"{export_format.upper()} analytics exports."
            ),
            code="analytics_export_dependency_missing",
            details={
                "dependency": dependency,
                "format": export_format,
            },
        )
        
class AnalyticsService:
    def __init__(
        self,
        repository: AnalyticsRepository | None = None,
    ) -> None:
        self.repository = repository or analytics_repository

    @staticmethod
    def utc_now() -> datetime:
        return datetime.now(tz=UTC)

    @staticmethod
    def ensure_timezone(value: datetime) -> datetime:
        if not isinstance(value, datetime):
            raise AnalyticsValidationError(
                "Expected a valid datetime value.",
                code="invalid_datetime",
                details={
                    "received_type": type(value).__name__,
                },
            )

        if value.tzinfo is None:
            return value.replace(tzinfo=UTC)

        return value.astimezone(UTC)

    @classmethod
    def start_of_day(cls, value: datetime) -> datetime:
        normalized = cls.ensure_timezone(value)

        return normalized.replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
        )

    @classmethod
    def end_of_day(cls, value: datetime) -> datetime:
        normalized = cls.ensure_timezone(value)

        return normalized.replace(
            hour=23,
            minute=59,
            second=59,
            microsecond=999_999,
        )

    @classmethod
    def start_of_month(cls, value: datetime) -> datetime:
        normalized = cls.ensure_timezone(value)

        return normalized.replace(
            day=1,
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
        )

    @classmethod
    def end_of_month(cls, value: datetime) -> datetime:
        normalized = cls.ensure_timezone(value)

        if normalized.month == 12:
            next_month = normalized.replace(
                year=normalized.year + 1,
                month=1,
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )
        else:
            next_month = normalized.replace(
                month=normalized.month + 1,
                day=1,
                hour=0,
                minute=0,
                second=0,
                microsecond=0,
            )

        return next_month - timedelta(microseconds=1)

    @classmethod
    def start_of_week(cls, value: datetime) -> datetime:
        normalized = cls.start_of_day(value)

        return normalized - timedelta(days=normalized.weekday())

    @classmethod
    def end_of_week(cls, value: datetime) -> datetime:
        return cls.start_of_week(value) + timedelta(
            days=6,
            hours=23,
            minutes=59,
            seconds=59,
            microseconds=999_999,
        )

    @classmethod
    def normalize_start_date(cls, value: datetime) -> datetime:
        return cls.start_of_day(value)

    @classmethod
    def normalize_end_date(cls, value: datetime) -> datetime:
        return cls.end_of_day(value)

    @classmethod
    def previous_period(
        cls,
        start_date: datetime,
        end_date: datetime,
    ) -> tuple[datetime, datetime]:
        normalized_start = cls.ensure_timezone(start_date)
        normalized_end = cls.ensure_timezone(end_date)

        if normalized_start > normalized_end:
            raise AnalyticsValidationError(
                "start_date must be earlier than or equal to end_date.",
                code="invalid_date_range",
                details={
                    "start_date": normalized_start.isoformat(),
                    "end_date": normalized_end.isoformat(),
                },
            )

        period_duration = normalized_end - normalized_start

        previous_end = normalized_start - timedelta(microseconds=1)
        previous_start = previous_end - period_duration

        return previous_start, previous_end

    @classmethod
    def default_date_range(
        cls,
        *,
        granularity: AnalyticsGranularity = "day",
        now: datetime | None = None,
    ) -> tuple[datetime, datetime]:
        current_time = cls.ensure_timezone(now or cls.utc_now())

        range_days = GRANULARITY_DEFAULT_RANGE_DAYS.get(
            granularity,
            DEFAULT_ANALYTICS_RANGE_DAYS,
        )

        end_date = cls.end_of_day(current_time)
        start_date = cls.start_of_day(
            end_date - timedelta(days=max(range_days - 1, 0))
        )

        return start_date, end_date

    @classmethod
    def normalize_optional_date_range(
        cls,
        *,
        start_date: datetime | None,
        end_date: datetime | None,
        granularity: AnalyticsGranularity = "day",
        default_days: int | None = None,
        now: datetime | None = None,
    ) -> tuple[datetime, datetime]:
        current_time = cls.ensure_timezone(now or cls.utc_now())

        if start_date is None and end_date is None:
            if default_days is None:
                return cls.default_date_range(
                    granularity=granularity,
                    now=current_time,
                )

            normalized_end = cls.end_of_day(current_time)
            normalized_start = cls.start_of_day(
                normalized_end
                - timedelta(days=max(default_days - 1, 0))
            )

            return normalized_start, normalized_end

        if start_date is None:
            normalized_end = cls.normalize_end_date(end_date)

            selected_days = (
                default_days
                if default_days is not None
                else GRANULARITY_DEFAULT_RANGE_DAYS.get(
                    granularity,
                    DEFAULT_ANALYTICS_RANGE_DAYS,
                )
            )

            normalized_start = cls.start_of_day(
                normalized_end
                - timedelta(days=max(selected_days - 1, 0))
            )

            return normalized_start, normalized_end

        if end_date is None:
            normalized_start = cls.normalize_start_date(start_date)
            normalized_end = cls.end_of_day(current_time)

            return normalized_start, normalized_end

        return (
            cls.normalize_start_date(start_date),
            cls.normalize_end_date(end_date),
        )

    @classmethod
    def create_date_range_schema(
        cls,
        *,
        start_date: datetime,
        end_date: datetime,
        granularity: AnalyticsGranularity,
    ) -> AnalyticsDateRangeSchema:
        return AnalyticsDateRangeSchema(
            start_date=cls.ensure_timezone(start_date),
            end_date=cls.ensure_timezone(end_date),
            granularity=granularity,
        )
        
    @staticmethod
    def validate_granularity(
        granularity: AnalyticsGranularity,
    ) -> AnalyticsGranularity:
        if granularity not in SUPPORTED_GRANULARITIES:
            raise AnalyticsValidationError(
                (
                    f"Unsupported analytics granularity '{granularity}'. "
                    f"Supported values are: "
                    f"{', '.join(SUPPORTED_GRANULARITIES)}."
                ),
                code="unsupported_granularity",
                details={
                    "granularity": granularity,
                    "supported_granularities": list(
                        SUPPORTED_GRANULARITIES
                    ),
                },
            )

        return granularity

    @staticmethod
    def validate_export_format(
        export_format: AnalyticsExportFormat,
    ) -> AnalyticsExportFormat:
        if export_format not in SUPPORTED_EXPORT_FORMATS:
            raise AnalyticsValidationError(
                (
                    f"Unsupported export format '{export_format}'. "
                    f"Supported values are: "
                    f"{', '.join(SUPPORTED_EXPORT_FORMATS)}."
                ),
                code="unsupported_export_format",
                details={
                    "format": export_format,
                    "supported_formats": list(
                        SUPPORTED_EXPORT_FORMATS
                    ),
                },
            )

        return export_format

    @staticmethod
    def validate_section(
        section: str,
    ) -> AnalyticsSection:
        normalized_section = str(section or "").strip().lower()

        if normalized_section not in SUPPORTED_ANALYTICS_SECTIONS:
            raise AnalyticsValidationError(
                (
                    f"Unsupported analytics section "
                    f"'{normalized_section or section}'. "
                    f"Supported values are: "
                    f"{', '.join(SUPPORTED_ANALYTICS_SECTIONS)}."
                ),
                code="unsupported_analytics_section",
                details={
                    "section": section,
                    "supported_sections": list(
                        SUPPORTED_ANALYTICS_SECTIONS
                    ),
                },
            )

        return normalized_section  # type: ignore[return-value]

    @staticmethod
    def validate_positive_integer(
        value: int,
        *,
        field_name: str,
        minimum: int = 1,
        maximum: int | None = None,
    ) -> int:
        if isinstance(value, bool) or not isinstance(value, int):
            raise AnalyticsValidationError(
                f"{field_name} must be an integer.",
                code="invalid_integer",
                details={
                    "field": field_name,
                    "value": value,
                },
            )

        if value < minimum:
            raise AnalyticsValidationError(
                f"{field_name} must be at least {minimum}.",
                code="integer_below_minimum",
                details={
                    "field": field_name,
                    "value": value,
                    "minimum": minimum,
                },
            )

        if maximum is not None and value > maximum:
            raise AnalyticsValidationError(
                f"{field_name} must not exceed {maximum}.",
                code="integer_above_maximum",
                details={
                    "field": field_name,
                    "value": value,
                    "maximum": maximum,
                },
            )

        return value

    @classmethod
    def validate_date_range(
        cls,
        *,
        start_date: datetime,
        end_date: datetime,
        granularity: AnalyticsGranularity,
        maximum_days: int | None = None,
    ) -> AnalyticsDateRangeSchema:
        validated_granularity = cls.validate_granularity(granularity)

        normalized_start = cls.normalize_start_date(start_date)
        normalized_end = cls.normalize_end_date(end_date)

        if normalized_start > normalized_end:
            raise AnalyticsValidationError(
                "start_date must be earlier than or equal to end_date.",
                code="invalid_date_range",
                details={
                    "start_date": normalized_start.isoformat(),
                    "end_date": normalized_end.isoformat(),
                },
            )

        allowed_days = (
            maximum_days
            if maximum_days is not None
            else GRANULARITY_MAX_RANGE_DAYS[
                validated_granularity
            ]
        )

        if allowed_days <= 0:
            raise AnalyticsValidationError(
                "maximum_days must be greater than zero.",
                code="invalid_maximum_days",
                details={
                    "maximum_days": allowed_days,
                },
            )

        range_duration = normalized_end - normalized_start
        range_days = range_duration.total_seconds() / 86_400

        if range_days > allowed_days:
            raise AnalyticsValidationError(
                (
                    f"The selected date range exceeds the maximum "
                    f"allowed range of {allowed_days} days for "
                    f"'{validated_granularity}' granularity."
                ),
                code="date_range_too_large",
                details={
                    "start_date": normalized_start.isoformat(),
                    "end_date": normalized_end.isoformat(),
                    "granularity": validated_granularity,
                    "maximum_days": allowed_days,
                    "selected_days": round(
                        range_days,
                        ROUNDING_PRECISION,
                    ),
                },
            )

        return AnalyticsDateRangeSchema(
            start_date=normalized_start,
            end_date=normalized_end,
            granularity=validated_granularity,
        )

    @classmethod
    def resolve_date_range(
        cls,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        granularity: AnalyticsGranularity = "day",
        default_days: int | None = None,
        maximum_days: int | None = None,
        now: datetime | None = None,
    ) -> AnalyticsDateRangeSchema:
        validated_granularity = cls.validate_granularity(granularity)

        if default_days is not None:
            cls.validate_positive_integer(
                default_days,
                field_name="default_days",
                minimum=1,
                maximum=MAX_ANALYTICS_RANGE_DAYS,
            )

        normalized_start, normalized_end = (
            cls.normalize_optional_date_range(
                start_date=start_date,
                end_date=end_date,
                granularity=validated_granularity,
                default_days=default_days,
                now=now,
            )
        )

        return cls.validate_date_range(
            start_date=normalized_start,
            end_date=normalized_end,
            granularity=validated_granularity,
            maximum_days=maximum_days,
        )

    @classmethod
    def resolve_export_date_range(
        cls,
        request: AnalyticsExportRequestSchema,
        *,
        now: datetime | None = None,
    ) -> AnalyticsDateRangeSchema:
        return cls.resolve_date_range(
            start_date=request.start_date,
            end_date=request.end_date,
            granularity=request.granularity,
            default_days=DEFAULT_EXPORT_RANGE_DAYS,
            maximum_days=MAX_ANALYTICS_RANGE_DAYS,
            now=now,
        )

    @classmethod
    def get_previous_period_from_range(
        cls,
        date_range: AnalyticsDateRangeSchema,
    ) -> tuple[datetime, datetime]:
        return cls.previous_period(
            date_range.start_date,
            date_range.end_date,
        )

    @classmethod
    def build_repository_date_context(
        cls,
        date_range: AnalyticsDateRangeSchema,
        *,
        now: datetime | None = None,
    ) -> dict[str, datetime]:
        current_time = cls.ensure_timezone(now or cls.utc_now())

        previous_start, previous_end = (
            cls.get_previous_period_from_range(date_range)
        )

        return {
            "start_date": date_range.start_date,
            "end_date": date_range.end_date,
            "previous_start_date": previous_start,
            "previous_end_date": previous_end,
            "today_start": cls.start_of_day(current_time),
            "month_start": cls.start_of_month(current_time),
        }

    @classmethod
    def validate_game_limit(
        cls,
        limit: int,
    ) -> int:
        return cls.validate_positive_integer(
            limit,
            field_name="top_limit",
            minimum=1,
            maximum=MAX_GAME_LIMIT,
        )
        
    @staticmethod
    def safe_int(
        value: Any,
        *,
        default: int = 0,
        minimum: int | None = None,
        maximum: int | None = None,
    ) -> int:
        if value is None:
            result = default
        elif isinstance(value, bool):
            result = int(value)
        elif isinstance(value, int):
            result = value
        elif isinstance(value, float):
            if not math.isfinite(value):
                result = default
            else:
                result = int(value)
        elif isinstance(value, str):
            normalized = value.strip().replace(",", "")

            if not normalized:
                result = default
            else:
                try:
                    result = int(float(normalized))
                except (TypeError, ValueError, OverflowError):
                    result = default
        else:
            try:
                result = int(value)
            except (TypeError, ValueError, OverflowError):
                result = default

        if minimum is not None:
            result = max(result, minimum)

        if maximum is not None:
            result = min(result, maximum)

        return result

    @staticmethod
    def safe_float(
        value: Any,
        *,
        default: float = 0.0,
        minimum: float | None = None,
        maximum: float | None = None,
        precision: int | None = None,
    ) -> float:
        if value is None:
            result = default
        elif isinstance(value, bool):
            result = float(value)
        elif isinstance(value, (int, float)):
            result = float(value)
        elif isinstance(value, str):
            normalized = value.strip().replace(",", "")

            if not normalized:
                result = default
            else:
                try:
                    result = float(normalized)
                except (TypeError, ValueError, OverflowError):
                    result = default
        else:
            try:
                result = float(value)
            except (TypeError, ValueError, OverflowError):
                result = default

        if not math.isfinite(result):
            result = default

        if minimum is not None:
            result = max(result, minimum)

        if maximum is not None:
            result = min(result, maximum)

        if precision is not None:
            result = round(result, precision)

        return result

    @classmethod
    def round_number(
        cls,
        value: Any,
        *,
        precision: int = ROUNDING_PRECISION,
        default: float = 0.0,
    ) -> float:
        return cls.safe_float(
            value,
            default=default,
            precision=precision,
        )

    @classmethod
    def round_money(
        cls,
        value: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        return cls.safe_float(
            value,
            default=0.0,
            minimum=0.0,
            precision=precision,
        )

    @classmethod
    def safe_percentage(
        cls,
        numerator: Any,
        denominator: Any,
        *,
        precision: int = ROUNDING_PRECISION,
        minimum: float = 0.0,
        maximum: float = 100.0,
    ) -> float:
        safe_numerator = cls.safe_float(numerator)
        safe_denominator = cls.safe_float(denominator)

        if safe_denominator == 0:
            return 0.0

        percentage = (
            safe_numerator
            / safe_denominator
        ) * 100

        return cls.safe_float(
            percentage,
            default=0.0,
            minimum=minimum,
            maximum=maximum,
            precision=precision,
        )

    @classmethod
    def safe_average(
        cls,
        total: Any,
        count: Any,
        *,
        precision: int = ROUNDING_PRECISION,
        minimum: float | None = 0.0,
    ) -> float:
        safe_total = cls.safe_float(total)
        safe_count = cls.safe_float(count)

        if safe_count <= 0:
            return 0.0

        return cls.safe_float(
            safe_total / safe_count,
            default=0.0,
            minimum=minimum,
            precision=precision,
        )

    @classmethod
    def safe_sum(
        cls,
        values: Sequence[Any] | None,
        *,
        precision: int | None = None,
    ) -> float:
        if not values:
            return 0.0

        total = sum(
            cls.safe_float(value)
            for value in values
        )

        if precision is not None:
            return round(total, precision)

        return total

    @classmethod
    def safe_count(
        cls,
        value: Any,
    ) -> int:
        return cls.safe_int(
            value,
            default=0,
            minimum=0,
        )

    @classmethod
    def safe_coin_amount(
        cls,
        value: Any,
    ) -> int:
        return cls.safe_int(
            value,
            default=0,
            minimum=0,
        )

    @classmethod
    def safe_signed_coin_amount(
        cls,
        value: Any,
    ) -> int:
        return cls.safe_int(
            value,
            default=0,
        )

    @classmethod
    def clamp(
        cls,
        value: Any,
        *,
        minimum: float,
        maximum: float,
        precision: int | None = None,
    ) -> float:
        if minimum > maximum:
            raise AnalyticsValidationError(
                "minimum cannot be greater than maximum.",
                code="invalid_clamp_range",
                details={
                    "minimum": minimum,
                    "maximum": maximum,
                },
            )

        return cls.safe_float(
            value,
            minimum=minimum,
            maximum=maximum,
            precision=precision,
        )

    @staticmethod
    def coalesce(
        *values: Any,
        default: Any = None,
    ) -> Any:
        for value in values:
            if value is not None:
                return value

        return default

    @staticmethod
    def normalize_label(
        value: Any,
        *,
        default: str = EMPTY_LABEL,
        max_length: int = 160,
    ) -> str:
        if value is None:
            return default

        normalized = str(value).strip()

        if not normalized:
            return default

        if len(normalized) > max_length:
            return normalized[:max_length]

        return normalized
    
    @classmethod
    def calculate_growth(
        cls,
        current_value: Any,
        previous_value: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        current = cls.safe_float(current_value)
        previous = cls.safe_float(previous_value)

        if previous == 0:
            if current == 0:
                return 0.0

            return 100.0 if current > 0 else -100.0

        growth = (
            (current - previous)
            / abs(previous)
        ) * 100

        return cls.safe_float(
            growth,
            default=0.0,
            precision=precision,
        )

    @classmethod
    def calculate_change(
        cls,
        current_value: Any,
        previous_value: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        current = cls.safe_float(current_value)
        previous = cls.safe_float(previous_value)

        return cls.safe_float(
            current - previous,
            default=0.0,
            precision=precision,
        )

    @classmethod
    def calculate_count_change(
        cls,
        current_value: Any,
        previous_value: Any,
    ) -> int:
        current = cls.safe_int(current_value)
        previous = cls.safe_int(previous_value)

        return current - previous

    @classmethod
    def growth_direction(
        cls,
        current_value: Any,
        previous_value: Any,
    ) -> GrowthDirection:
        current = cls.safe_float(current_value)
        previous = cls.safe_float(previous_value)

        if current > previous:
            return "increase"

        if current < previous:
            return "decrease"

        return "unchanged"

    @classmethod
    def add_growth(
        cls,
        current_value: Any,
        previous_value: Any,
        *,
        count_metric: bool = False,
        precision: int = ROUNDING_PRECISION,
    ) -> dict[str, int | float]:
        if count_metric:
            current = cls.safe_count(current_value)
            previous = cls.safe_count(previous_value)

            return {
                "current_value": current,
                "previous_value": previous,
                "change": current - previous,
                "growth_percentage": cls.calculate_growth(
                    current,
                    previous,
                    precision=precision,
                ),
            }

        current = cls.safe_float(
            current_value,
            precision=precision,
        )
        previous = cls.safe_float(
            previous_value,
            precision=precision,
        )

        return {
            "current_value": current,
            "previous_value": previous,
            "change": cls.calculate_change(
                current,
                previous,
                precision=precision,
            ),
            "growth_percentage": cls.calculate_growth(
                current,
                previous,
                precision=precision,
            ),
        }

    @classmethod
    def add_growth_fields(
        cls,
        target: Mapping[str, Any] | None,
        *,
        current_value: Any,
        previous_value: Any,
        prefix: str = "",
        precision: int = ROUNDING_PRECISION,
    ) -> dict[str, Any]:
        result = dict(target or {})

        normalized_prefix = str(prefix or "").strip()

        if normalized_prefix:
            normalized_prefix = f"{normalized_prefix}_"

        result.update(
            {
                f"{normalized_prefix}current_value": cls.safe_float(
                    current_value,
                    precision=precision,
                ),
                f"{normalized_prefix}previous_value": cls.safe_float(
                    previous_value,
                    precision=precision,
                ),
                f"{normalized_prefix}change": cls.calculate_change(
                    current_value,
                    previous_value,
                    precision=precision,
                ),
                f"{normalized_prefix}growth_percentage": (
                    cls.calculate_growth(
                        current_value,
                        previous_value,
                        precision=precision,
                    )
                ),
                f"{normalized_prefix}growth_direction": (
                    cls.growth_direction(
                        current_value,
                        previous_value,
                    )
                ),
            }
        )

        return result

    @classmethod
    def compare_period_values(
        cls,
        *,
        current: Mapping[str, Any] | None,
        previous: Mapping[str, Any] | None,
        fields: Sequence[str],
        count_fields: Sequence[str] | None = None,
        precision: int = ROUNDING_PRECISION,
    ) -> dict[str, dict[str, int | float]]:
        current_data = dict(current or {})
        previous_data = dict(previous or {})
        count_field_set = set(count_fields or [])

        comparison: dict[str, dict[str, int | float]] = {}

        for field_name in fields:
            comparison[field_name] = cls.add_growth(
                current_data.get(field_name),
                previous_data.get(field_name),
                count_metric=field_name in count_field_set,
                precision=precision,
            )

        return comparison

    @classmethod
    def calculate_retention_rate(
        cls,
        returning_users: Any,
        total_users: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        return cls.safe_percentage(
            returning_users,
            total_users,
            precision=precision,
        )

    @classmethod
    def calculate_conversion_rate(
        cls,
        converted_count: Any,
        eligible_count: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        return cls.safe_percentage(
            converted_count,
            eligible_count,
            precision=precision,
        )

    @classmethod
    def calculate_net_flow(
        cls,
        credited_amount: Any,
        debited_amount: Any,
    ) -> int:
        credited = cls.safe_coin_amount(credited_amount)
        debited = cls.safe_coin_amount(debited_amount)

        return credited - debited

    @classmethod
    def calculate_average_revenue_per_user(
        cls,
        revenue: Any,
        user_count: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        return cls.safe_average(
            revenue,
            user_count,
            precision=precision,
            minimum=0.0,
        )

    @classmethod
    def calculate_percentage_share(
        cls,
        value: Any,
        total: Any,
        *,
        precision: int = ROUNDING_PRECISION,
    ) -> float:
        return cls.safe_percentage(
            value,
            total,
            precision=precision,
        )
        
    @classmethod
    def format_chart(
        cls,
        rows: Sequence[Mapping[str, Any]] | None,
        *,
        label_keys: Sequence[str] = ("label", "_id"),
        date_keys: Sequence[str] = ("date", "_id"),
        value_keys: Sequence[str] = ("value", "total", "amount"),
        count: bool = False,
    ) -> list[
        AnalyticsChartPointSchema
        | AnalyticsCountChartPointSchema
    ]:
        if not rows:
            return []

        formatted: list[
            AnalyticsChartPointSchema
            | AnalyticsCountChartPointSchema
        ] = []

        for row in rows:
            row_data = dict(row or {})

            raw_date = next(
                (
                    row_data.get(key)
                    for key in date_keys
                    if row_data.get(key) is not None
                ),
                None,
            )

            if not isinstance(raw_date, datetime):
                continue

            normalized_date = cls.ensure_timezone(raw_date)

            raw_label = next(
                (
                    row_data.get(key)
                    for key in label_keys
                    if row_data.get(key) not in (None, "")
                ),
                None,
            )

            label = cls.normalize_label(
                raw_label,
                default=normalized_date.isoformat(),
                max_length=120,
            )

            raw_value = next(
                (
                    row_data.get(key)
                    for key in value_keys
                    if row_data.get(key) is not None
                ),
                0,
            )

            if count:
                formatted.append(
                    AnalyticsCountChartPointSchema(
                        label=label,
                        date=normalized_date,
                        value=cls.safe_count(raw_value),
                    )
                )
            else:
                formatted.append(
                    AnalyticsChartPointSchema(
                        label=label,
                        date=normalized_date,
                        value=cls.safe_float(
                            raw_value,
                            precision=ROUNDING_PRECISION,
                        ),
                    )
                )

        return sorted(
            formatted,
            key=lambda item: item.date,
        )

    @classmethod
    def format_distribution(
        cls,
        rows: Sequence[Mapping[str, Any]] | None,
        *,
        label_keys: Sequence[str] = (
            "label",
            "name",
            "type",
            "status",
            "category",
            "_id",
        ),
        value_keys: Sequence[str] = (
            "value",
            "count",
            "total",
            "amount",
        ),
        total: Any | None = None,
        count: bool = False,
    ) -> list[
        AnalyticsCategoryValueSchema
        | AnalyticsCountCategorySchema
    ]:
        if not rows:
            return []

        normalized_rows: list[tuple[str, int | float]] = []

        for row in rows:
            row_data = dict(row or {})

            raw_label = next(
                (
                    row_data.get(key)
                    for key in label_keys
                    if row_data.get(key) not in (None, "")
                ),
                EMPTY_LABEL,
            )

            raw_value = next(
                (
                    row_data.get(key)
                    for key in value_keys
                    if row_data.get(key) is not None
                ),
                0,
            )

            label = cls.normalize_label(raw_label)

            value: int | float

            if count:
                value = cls.safe_count(raw_value)
            else:
                value = cls.safe_float(
                    raw_value,
                    precision=ROUNDING_PRECISION,
                )

            normalized_rows.append((label, value))

        calculated_total = (
            cls.safe_float(total)
            if total is not None
            else sum(
                cls.safe_float(value)
                for _, value in normalized_rows
            )
        )

        formatted: list[
            AnalyticsCategoryValueSchema
            | AnalyticsCountCategorySchema
        ] = []

        for label, value in normalized_rows:
            percentage = cls.safe_percentage(
                value,
                calculated_total,
            )

            if count:
                formatted.append(
                    AnalyticsCountCategorySchema(
                        label=label,
                        value=cls.safe_count(value),
                        percentage=percentage,
                    )
                )
            else:
                formatted.append(
                    AnalyticsCategoryValueSchema(
                        label=label,
                        value=cls.safe_float(
                            value,
                            precision=ROUNDING_PRECISION,
                        ),
                        percentage=percentage,
                    )
                )

        return sorted(
            formatted,
            key=lambda item: item.value,
            reverse=True,
        )

    @classmethod
    def format_game_table(
        cls,
        rows: Sequence[Mapping[str, Any]] | None,
        *,
        total_plays: Any = 0,
        limit: int = DEFAULT_GAME_LIMIT,
    ) -> list[GamePerformanceItemSchema]:
        if not rows:
            return []

        validated_limit = cls.validate_game_limit(limit)
        safe_total_plays = cls.safe_coin_amount(total_plays)

        formatted: list[GamePerformanceItemSchema] = []

        for row in rows[:validated_limit]:
            row_data = dict(row or {})

            raw_game_id = cls.coalesce(
                row_data.get("game_id"),
                row_data.get("_id"),
                row_data.get("id"),
            )

            game_id = cls.normalize_label(
                raw_game_id,
                default="unknown",
                max_length=200,
            )

            name = cls.normalize_label(
                cls.coalesce(
                    row_data.get("name"),
                    row_data.get("title"),
                ),
                default="Unknown Game",
                max_length=120,
            )

            slug = cls.normalize_label(
                row_data.get("slug"),
                default=game_id,
                max_length=140,
            )

            play_count = cls.safe_count(
                cls.coalesce(
                    row_data.get("play_count"),
                    row_data.get("plays"),
                    row_data.get("total_plays"),
                    default=0,
                )
            )

            formatted.append(
                GamePerformanceItemSchema(
                    game_id=game_id,
                    name=name,
                    slug=slug,
                    category=cls.normalize_label(
                        row_data.get("category"),
                        default="other",
                        max_length=80,
                    ),
                    provider_name=(
                        cls.normalize_label(
                            cls.coalesce(
                                row_data.get("provider_name"),
                                row_data.get("provider"),
                            ),
                            max_length=120,
                        )
                        if cls.coalesce(
                            row_data.get("provider_name"),
                            row_data.get("provider"),
                        )
                        not in (None, "")
                        else None
                    ),
                    status=cls.normalize_label(
                        row_data.get("status"),
                        default="draft",
                        max_length=40,
                    ),
                    is_featured=bool(
                        row_data.get("is_featured", False)
                    ),
                    show_on_landing_page=bool(
                        row_data.get(
                            "show_on_landing_page",
                            False,
                        )
                    ),
                    play_count=play_count,
                    percentage_of_total_plays=(
                        cls.safe_percentage(
                            play_count,
                            safe_total_plays,
                        )
                    ),
                )
            )

        return formatted

    @staticmethod
    def mapping_value(
        source: Mapping[str, Any] | None,
        *keys: str,
        default: Any = None,
    ) -> Any:
        if not source:
            return default

        for key in keys:
            if key in source and source[key] is not None:
                return source[key]

        return default

    @classmethod
    def normalize_mapping(
        cls,
        value: Any,
    ) -> dict[str, Any]:
        if isinstance(value, Mapping):
            return dict(value)

        return {}

    @classmethod
    def normalize_sequence(
        cls,
        value: Any,
    ) -> list[Mapping[str, Any]]:
        if not isinstance(value, Sequence):
            return []

        if isinstance(value, (str, bytes, bytearray)):
            return []

        return [
            dict(item)
            for item in value
            if isinstance(item, Mapping)
        ]

    @classmethod
    def serialize_value(
        cls,
        value: Any,
    ) -> Any:
        if isinstance(value, datetime):
            return cls.ensure_timezone(value).isoformat()

        if hasattr(value, "model_dump"):
            return cls.serialize_value(
                value.model_dump(mode="python")
            )

        if isinstance(value, Mapping):
            return {
                str(key): cls.serialize_value(item)
                for key, item in value.items()
            }

        if isinstance(value, Sequence) and not isinstance(
            value,
            (str, bytes, bytearray),
        ):
            return [
                cls.serialize_value(item)
                for item in value
            ]

        if isinstance(value, float):
            if not math.isfinite(value):
                return 0.0

            return round(value, ROUNDING_PRECISION)

        return value

    @classmethod
    def flatten_mapping(
        cls,
        value: Mapping[str, Any],
        *,
        prefix: str = "",
        separator: str = ".",
    ) -> dict[str, Any]:
        flattened: dict[str, Any] = {}

        for key, item in value.items():
            normalized_key = str(key)
            full_key = (
                f"{prefix}{separator}{normalized_key}"
                if prefix
                else normalized_key
            )

            if isinstance(item, Mapping):
                flattened.update(
                    cls.flatten_mapping(
                        item,
                        prefix=full_key,
                        separator=separator,
                    )
                )
            elif isinstance(item, Sequence) and not isinstance(
                item,
                (str, bytes, bytearray),
            ):
                flattened[full_key] = json.dumps(
                    cls.serialize_value(item),
                    ensure_ascii=False,
                )
            else:
                flattened[full_key] = (
                    cls.serialize_value(item)
                )

        return flattened

    @staticmethod
    def create_buffer(
        initial_bytes: bytes | None = None,
    ) -> BytesIO:
        buffer = BytesIO(initial_bytes or b"")
        buffer.seek(0)
        return buffer

    @staticmethod
    def create_text_buffer() -> StringIO:
        return StringIO(
            newline=CSV_NEWLINE,
        )

    @classmethod
    def get_export_content_type(
        cls,
        export_format: AnalyticsExportFormat,
    ) -> str:
        validated_format = cls.validate_export_format(
            export_format
        )

        return EXPORT_CONTENT_TYPES[validated_format]

    @classmethod
    def get_export_extension(
        cls,
        export_format: AnalyticsExportFormat,
    ) -> str:
        validated_format = cls.validate_export_format(
            export_format
        )

        return EXPORT_FILE_EXTENSIONS[validated_format]

    @classmethod
    def build_export_filename(
        cls,
        *,
        section: AnalyticsSection,
        export_format: AnalyticsExportFormat,
        generated_at: datetime | None = None,
    ) -> str:
        validated_section = cls.validate_section(section)
        validated_format = cls.validate_export_format(
            export_format
        )

        timestamp = cls.ensure_timezone(
            generated_at or cls.utc_now()
        ).strftime("%Y%m%d_%H%M%S")

        extension = cls.get_export_extension(
            validated_format
        )

        return (
            f"{EXPORT_FILENAME_PREFIX}_"
            f"{validated_section}_"
            f"{timestamp}.{extension}"
        )

    @classmethod
    def build_export_response(
        cls,
        *,
        export_format: AnalyticsExportFormat,
        section: AnalyticsSection,
        filename: str,
        generated_at: datetime | None = None,
    ) -> AnalyticsExportResponseSchema:
        validated_format = cls.validate_export_format(
            export_format
        )
        validated_section = cls.validate_section(section)

        normalized_filename = cls.normalize_label(
            filename,
            default=cls.build_export_filename(
                section=validated_section,
                export_format=validated_format,
                generated_at=generated_at,
            ),
            max_length=255,
        )

        return AnalyticsExportResponseSchema(
            success=True,
            format=validated_format,
            section=validated_section,
            filename=normalized_filename,
            content_type=cls.get_export_content_type(
                validated_format
            ),
            generated_at=cls.ensure_timezone(
                generated_at or cls.utc_now()
            ),
        )

    @classmethod
    def serialize_export_payload(
        cls,
        payload: Any,
    ) -> dict[str, Any] | list[Any] | Any:
        if hasattr(payload, "model_dump"):
            payload = payload.model_dump(
                mode="python",
                exclude_none=False,
            )

        return cls.serialize_value(payload)

    @classmethod
    def payload_to_flat_rows(
        cls,
        payload: Any,
        *,
        section: str = "analytics",
    ) -> list[dict[str, Any]]:
        serialized = cls.serialize_export_payload(payload)

        if isinstance(serialized, Mapping):
            rows: list[dict[str, Any]] = []

            scalar_values: dict[str, Any] = {}
            nested_values: dict[str, Any] = {}

            for key, value in serialized.items():
                if isinstance(value, Mapping):
                    nested_values[str(key)] = value
                elif isinstance(value, list):
                    nested_values[str(key)] = value
                else:
                    scalar_values[str(key)] = value

            if scalar_values:
                rows.append(
                    {
                        "section": section,
                        **scalar_values,
                    }
                )

            for key, value in nested_values.items():
                nested_section = (
                    f"{section}.{key}"
                    if section
                    else key
                )

                rows.extend(
                    cls.payload_to_flat_rows(
                        value,
                        section=nested_section,
                    )
                )

            if not rows:
                rows.append(
                    {
                        "section": section,
                    }
                )

            return rows

        if isinstance(serialized, list):
            rows = []

            if not serialized:
                return [
                    {
                        "section": section,
                    }
                ]

            for index, item in enumerate(serialized):
                if isinstance(item, Mapping):
                    flattened = cls.flatten_mapping(item)

                    rows.append(
                        {
                            "section": section,
                            "row": index + 1,
                            **flattened,
                        }
                    )
                else:
                    rows.append(
                        {
                            "section": section,
                            "row": index + 1,
                            "value": item,
                        }
                    )

            return rows

        return [
            {
                "section": section,
                "value": serialized,
            }
        ]

    @classmethod
    def collect_export_headers(
        cls,
        rows: Sequence[Mapping[str, Any]],
    ) -> list[str]:
        headers: list[str] = []
        seen: set[str] = set()

        preferred_headers = (
            "section",
            "row",
            "label",
            "date",
            "value",
            "percentage",
        )

        for preferred_header in preferred_headers:
            if any(
                preferred_header in row
                for row in rows
            ):
                headers.append(preferred_header)
                seen.add(preferred_header)

        for row in rows:
            for key in row.keys():
                normalized_key = str(key)

                if normalized_key not in seen:
                    seen.add(normalized_key)
                    headers.append(normalized_key)

        return headers

    @classmethod
    def normalize_export_cell(
        cls,
        value: Any,
    ) -> str | int | float | bool:
        serialized = cls.serialize_value(value)

        if serialized is None:
            return ""

        if isinstance(serialized, (str, int, float, bool)):
            return serialized

        return json.dumps(
            serialized,
            ensure_ascii=False,
            separators=(",", ":"),
        )

    @classmethod
    def write_csv_rows(
        cls,
        rows: Sequence[Mapping[str, Any]],
    ) -> bytes:
        normalized_rows = [
            dict(row)
            for row in rows
        ]

        if not normalized_rows:
            normalized_rows = [
                {
                    "section": "analytics",
                    "value": "",
                }
            ]

        headers = cls.collect_export_headers(
            normalized_rows
        )

        text_buffer = cls.create_text_buffer()

        writer = csv.DictWriter(
            text_buffer,
            fieldnames=headers,
            extrasaction="ignore",
        )

        writer.writeheader()

        for row in normalized_rows:
            writer.writerow(
                {
                    header: cls.normalize_export_cell(
                        row.get(header)
                    )
                    for header in headers
                }
            )

        content = text_buffer.getvalue()
        text_buffer.close()

        return content.encode(CSV_ENCODING)

    @classmethod
    def generate_csv_buffer(
        cls,
        payload: Any,
        *,
        section: str,
    ) -> BytesIO:
        try:
            rows = cls.payload_to_flat_rows(
                payload,
                section=section,
            )

            csv_bytes = cls.write_csv_rows(rows)

            return cls.create_buffer(csv_bytes)

        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsExportError(
                "Unable to generate the CSV analytics export.",
                code="csv_export_failed",
                details={
                    "section": section,
                    "error": str(exc),
                },
            ) from exc

    @staticmethod
    def rewind_buffer(
        buffer: BytesIO,
    ) -> BytesIO:
        buffer.seek(0)
        return buffer

    @staticmethod
    def buffer_size(
        buffer: BytesIO,
    ) -> int:
        current_position = buffer.tell()

        buffer.seek(0, 2)
        size = buffer.tell()
        buffer.seek(current_position)

        return size

    @classmethod
    def validate_export_buffer(
        cls,
        buffer: BytesIO,
        *,
        export_format: AnalyticsExportFormat,
    ) -> BytesIO:
        if not isinstance(buffer, BytesIO):
            raise AnalyticsExportError(
                "The generated export is not a valid binary buffer.",
                code="invalid_export_buffer",
                details={
                    "format": export_format,
                    "received_type": type(buffer).__name__,
                },
            )

        if cls.buffer_size(buffer) <= 0:
            raise AnalyticsExportError(
                "The generated analytics export is empty.",
                code="empty_export_buffer",
                details={
                    "format": export_format,
                },
            )

        return cls.rewind_buffer(buffer)
    
    async def get_overview(
        self,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
    ) -> AnalyticsOverviewResponseSchema:
        date_range = self.resolve_date_range(
            start_date=start_date,
            end_date=end_date,
            default_days=DEFAULT_ANALYTICS_RANGE_DAYS,
        )

        repository_context = self.build_repository_date_context(
            date_range
        )

        try:
            repository_result = await self.repository.get_overview(
                start_date=repository_context["start_date"],
                end_date=repository_context["end_date"],
                previous_start_date=repository_context[
                    "previous_start_date"
                ],
                previous_end_date=repository_context[
                    "previous_end_date"
                ],
                today_start=repository_context["today_start"],
                month_start=repository_context["month_start"],
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsRepositoryError(
                "Unable to retrieve overview analytics.",
                code="overview_repository_failed",
                details={
                    "start_date": date_range.start_date.isoformat(),
                    "end_date": date_range.end_date.isoformat(),
                    "error": str(exc),
                },
            ) from exc

        repository_data = self.normalize_mapping(
            repository_result
        )

        users_data = self.normalize_mapping(
            repository_data.get("users")
        )
        wallet_data = self.normalize_mapping(
            repository_data.get("wallet")
        )
        transaction_data = self.normalize_mapping(
            repository_data.get("transactions")
        )
        game_data = self.normalize_mapping(
            repository_data.get("games")
        )
        recharge_data = self.normalize_mapping(
            repository_data.get("recharge")
        )

        overview_users = self.build_overview_users(
            users_data
        )
        overview_wallet = self.build_overview_wallet(
            wallet_data
        )
        overview_transactions = (
            self.build_overview_transactions(
                transaction_data
            )
        )
        overview_games = self.build_overview_games(
            game_data
        )
        overview_recharge = self.build_overview_recharge(
            recharge_data
        )
        overview_kpis = self.build_overview_kpis(
            users=users_data,
            wallet=wallet_data,
            transactions=transaction_data,
        )

        try:
            return AnalyticsOverviewResponseSchema(
                range=date_range,
                generated_at=self.utc_now(),
                users=overview_users,
                wallet=overview_wallet,
                transactions=overview_transactions,
                games=overview_games,
                recharge=overview_recharge,
                kpis=overview_kpis,
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsServiceError(
                "Unable to build the overview analytics response.",
                code="overview_response_failed",
                details={
                    "start_date": date_range.start_date.isoformat(),
                    "end_date": date_range.end_date.isoformat(),
                    "error": str(exc),
                },
            ) from exc

    @classmethod
    def build_overview_users(
        cls,
        source: Mapping[str, Any] | None,
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)

        current_new_users = cls.safe_count(
            data.get("new_users_in_range")
        )
        previous_new_users = cls.safe_count(
            data.get("previous_period_new_users")
        )

        growth_percentage = cls.calculate_growth(
            current_new_users,
            previous_new_users,
        )

        return {
            "total_users": cls.safe_count(
                data.get("total_users")
            ),
            "total_players": cls.safe_count(
                data.get("total_players")
            ),
            "active_users": cls.safe_count(
                data.get("active_users")
            ),
            "pending_users": cls.safe_count(
                data.get("pending_users")
            ),
            "suspended_users": cls.safe_count(
                data.get("suspended_users")
            ),
            "blocked_users": cls.safe_count(
                data.get("blocked_users")
            ),
            "verified_users": cls.safe_count(
                data.get("verified_users")
            ),
            "unverified_users": cls.safe_count(
                data.get("unverified_users")
            ),
            "new_users_today": cls.safe_count(
                data.get("new_users_today")
            ),
            "new_users_in_range": current_new_users,
            "returning_users": cls.safe_count(
                data.get("returning_users")
            ),
            "growth_percentage": growth_percentage,
        }

    @classmethod
    def build_overview_wallet(
        cls,
        source: Mapping[str, Any] | None,
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)

        return {
            "total_wallets": cls.safe_count(
                data.get("total_wallets")
            ),
            "active_wallets": cls.safe_count(
                data.get("active_wallets")
            ),
            "frozen_wallets": cls.safe_count(
                data.get("frozen_wallets")
            ),
            "zero_balance_wallets": cls.safe_count(
                data.get("zero_balance_wallets")
            ),
            "positive_balance_wallets": cls.safe_count(
                data.get("positive_balance_wallets")
            ),
            "total_coins_in_circulation": (
                cls.safe_coin_amount(
                    data.get("total_coins_in_circulation")
                )
            ),
            "average_wallet_balance": cls.round_money(
                data.get("average_wallet_balance")
            ),
        }

    @classmethod
    def build_overview_transactions(
        cls,
        source: Mapping[str, Any] | None,
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)

        credited_coins = cls.safe_coin_amount(
            data.get("total_credited_coins")
        )
        debited_coins = cls.safe_coin_amount(
            data.get("total_debited_coins")
        )

        repository_net_flow = data.get("net_coin_flow")

        net_coin_flow = (
            cls.safe_signed_coin_amount(
                repository_net_flow
            )
            if repository_net_flow is not None
            else cls.calculate_net_flow(
                credited_coins,
                debited_coins,
            )
        )

        return {
            "total_transactions": cls.safe_count(
                data.get("total_transactions")
            ),
            "purchase_transactions": cls.safe_count(
                data.get("purchase_transactions")
            ),
            "game_entry_transactions": cls.safe_count(
                data.get("game_entry_transactions")
            ),
            "admin_credit_transactions": cls.safe_count(
                data.get("admin_credit_transactions")
            ),
            "admin_debit_transactions": cls.safe_count(
                data.get("admin_debit_transactions")
            ),
            "refund_transactions": cls.safe_count(
                data.get("refund_transactions")
            ),
            "total_credited_coins": credited_coins,
            "total_debited_coins": debited_coins,
            "net_coin_flow": net_coin_flow,
            "average_transaction_amount": cls.round_money(
                data.get("average_transaction_amount")
            ),
            "transactions_today": cls.safe_count(
                data.get("transactions_today")
            ),
        }

    @classmethod
    def build_overview_games(
        cls,
        source: Mapping[str, Any] | None,
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)

        return {
            "total_games": cls.safe_count(
                data.get("total_games")
            ),
            "published_games": cls.safe_count(
                data.get("published_games")
            ),
            "draft_games": cls.safe_count(
                data.get("draft_games")
            ),
            "maintenance_games": cls.safe_count(
                data.get("maintenance_games")
            ),
            "disabled_games": cls.safe_count(
                data.get("disabled_games")
            ),
            "featured_games": cls.safe_count(
                data.get("featured_games")
            ),
            "landing_page_games": cls.safe_count(
                data.get("landing_page_games")
            ),
            "total_play_count": cls.safe_count(
                data.get("total_play_count")
            ),
            "games_played_today": cls.safe_count(
                data.get("games_played_today")
            ),
        }

    @classmethod
    def build_overview_recharge(
        cls,
        source: Mapping[str, Any] | None,
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)

        return {
            "total_packages": cls.safe_count(
                data.get("total_packages")
            ),
            "active_packages": cls.safe_count(
                data.get("active_packages")
            ),
            "inactive_packages": cls.safe_count(
                data.get("inactive_packages")
            ),
            "lowest_price": cls.round_money(
                data.get("lowest_price")
            ),
            "highest_price": cls.round_money(
                data.get("highest_price")
            ),
            "total_base_coins": cls.safe_coin_amount(
                data.get("total_base_coins")
            ),
            "total_bonus_coins": cls.safe_coin_amount(
                data.get("total_bonus_coins")
            ),
        }

    @classmethod
    def build_overview_kpis(
        cls,
        *,
        users: Mapping[str, Any] | None,
        wallet: Mapping[str, Any] | None,
        transactions: Mapping[str, Any] | None,
    ) -> dict[str, Any]:
        user_data = cls.normalize_mapping(users)
        wallet_data = cls.normalize_mapping(wallet)
        transaction_data = cls.normalize_mapping(
            transactions
        )

        total_users = cls.safe_count(
            user_data.get("total_users")
        )
        active_users = cls.safe_count(
            user_data.get("active_users")
        )
        verified_users = cls.safe_count(
            user_data.get("verified_users")
        )
        daily_active_users = cls.safe_count(
            user_data.get("daily_active_users")
        )
        monthly_active_users = cls.safe_count(
            user_data.get("monthly_active_users")
        )

        total_wallets = cls.safe_count(
            wallet_data.get("total_wallets")
        )
        credited_coins = cls.safe_coin_amount(
            transaction_data.get("total_credited_coins")
        )
        debited_coins = cls.safe_coin_amount(
            transaction_data.get("total_debited_coins")
        )

        repository_net_flow = transaction_data.get(
            "net_coin_flow"
        )

        net_coin_flow = (
            cls.safe_signed_coin_amount(
                repository_net_flow
            )
            if repository_net_flow is not None
            else cls.calculate_net_flow(
                credited_coins,
                debited_coins,
            )
        )

        return {
            "daily_active_users": daily_active_users,
            "monthly_active_users": monthly_active_users,
            "average_revenue_per_user": (
                cls.calculate_average_revenue_per_user(
                    net_coin_flow,
                    total_users,
                )
            ),
            "wallet_conversion_rate": (
                cls.calculate_conversion_rate(
                    total_wallets,
                    total_users,
                )
            ),
            "verified_user_rate": cls.safe_percentage(
                verified_users,
                total_users,
            ),
            "active_user_rate": cls.safe_percentage(
                active_users,
                total_users,
            ),
        }

    async def get_revenue(
        self,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        granularity: AnalyticsGranularity = "day",
    ) -> RevenueAnalyticsResponseSchema:
        date_range = self.resolve_date_range(
            start_date=start_date,
            end_date=end_date,
            granularity=granularity,
        )
        context = self.build_repository_date_context(date_range)
        try:
            result = await self.repository.get_revenue_analytics(
                start_date=context["start_date"],
                end_date=context["end_date"],
                previous_start_date=context["previous_start_date"],
                previous_end_date=context["previous_end_date"],
                granularity=date_range.granularity,
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsRepositoryError(
                "Unable to retrieve revenue analytics.",
                code="revenue_repository_failed",
                details={"error": str(exc)},
            ) from exc

        data = self.normalize_mapping(result)
        summary_data = self.normalize_mapping(data.get("summary"))
        total_users = await self._get_total_user_count_for_arpu(
            date_range=date_range
        )
        summary = self.build_revenue_summary(
            summary_data,
            total_users=total_users,
        )
        return RevenueAnalyticsResponseSchema(
            range=date_range,
            generated_at=self.utc_now(),
            summary=summary,
            revenue_trend=self.format_chart(
                self.normalize_sequence(data.get("revenue_trend"))
            ),
            credit_trend=self.format_chart(
                self.normalize_sequence(data.get("credit_trend"))
            ),
            debit_trend=self.format_chart(
                self.normalize_sequence(data.get("debit_trend"))
            ),
            revenue_by_transaction_type=self.format_distribution(
                self.normalize_sequence(
                    data.get("revenue_by_transaction_type")
                ),
                count=False,
            ),
        )

    async def _get_total_user_count_for_arpu(
        self,
        *,
        date_range: AnalyticsDateRangeSchema,
    ) -> int:
        try:
            context = self.build_repository_date_context(date_range)
            user_summary = await self.repository.get_user_summary(
                start_date=date_range.start_date,
                end_date=date_range.end_date,
                today_start=context["today_start"],
                month_start=context["month_start"],
            )
            return self.safe_count(
                self.normalize_mapping(user_summary).get("total_users")
            )
        except Exception:
            return 0

    @classmethod
    def build_revenue_summary(
        cls,
        source: Mapping[str, Any] | None,
        *,
        total_users: Any = 0,
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)
        credited = cls.safe_coin_amount(data.get("total_credited_coins"))
        debited = cls.safe_coin_amount(data.get("total_debited_coins"))
        net = cls.safe_signed_coin_amount(
            data.get("net_coin_flow")
        ) if data.get("net_coin_flow") is not None else cls.calculate_net_flow(
            credited, debited
        )
        return {
            "total_credited_coins": credited,
            "total_debited_coins": debited,
            "purchase_coins": cls.safe_coin_amount(data.get("purchase_coins")),
            "refund_coins": cls.safe_coin_amount(data.get("refund_coins")),
            "game_entry_coins": cls.safe_coin_amount(data.get("game_entry_coins")),
            "admin_credit_coins": cls.safe_coin_amount(data.get("admin_credit_coins")),
            "admin_debit_coins": cls.safe_coin_amount(data.get("admin_debit_coins")),
            "net_coin_flow": net,
            "average_credit_amount": cls.round_money(data.get("average_credit_amount")),
            "average_debit_amount": cls.round_money(data.get("average_debit_amount")),
            "average_revenue_per_user": cls.calculate_average_revenue_per_user(
                max(net, 0), total_users
            ),
            "growth_percentage": cls.calculate_growth(
                net, data.get("previous_net_coin_flow")
            ),
        }

    async def get_users(
        self,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        granularity: AnalyticsGranularity = "day",
    ) -> UserAnalyticsResponseSchema:
        date_range = self.resolve_date_range(
            start_date=start_date,
            end_date=end_date,
            granularity=granularity,
        )
        context = self.build_repository_date_context(date_range)
        try:
            result = await self.repository.get_user_analytics(
                start_date=context["start_date"],
                end_date=context["end_date"],
                previous_start_date=context["previous_start_date"],
                previous_end_date=context["previous_end_date"],
                today_start=context["today_start"],
                month_start=context["month_start"],
                granularity=date_range.granularity,
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsRepositoryError(
                "Unable to retrieve user analytics.",
                code="users_repository_failed",
                details={"error": str(exc)},
            ) from exc
        data = self.normalize_mapping(result)
        summary_data = self.normalize_mapping(data.get("summary"))
        return UserAnalyticsResponseSchema(
            range=date_range,
            generated_at=self.utc_now(),
            summary=self.build_user_summary(summary_data),
            registration_growth=self.format_chart(
                self.normalize_sequence(data.get("registration_growth")),
                count=True,
            ),
            users_by_role=self.format_distribution(
                self.normalize_sequence(data.get("users_by_role")), count=True
            ),
            users_by_status=self.format_distribution(
                self.normalize_sequence(data.get("users_by_status")), count=True
            ),
            users_by_country=self.format_distribution(
                self.normalize_sequence(data.get("users_by_country")), count=True
            ),
            users_by_language=self.format_distribution(
                self.normalize_sequence(data.get("users_by_language")), count=True
            ),
            verification_distribution=self.format_distribution(
                self.normalize_sequence(data.get("verification_distribution")),
                count=True,
            ),
        )

    @classmethod
    def build_user_summary(cls, source: Mapping[str, Any] | None) -> dict[str, Any]:
        data = cls.normalize_mapping(source)
        total = cls.safe_count(data.get("total_users"))
        active = cls.safe_count(data.get("active_users"))
        verified = cls.safe_count(data.get("verified_users"))
        current_new = cls.safe_count(data.get("new_users_in_range"))
        previous_new = cls.safe_count(data.get("previous_period_new_users"))
        return {
            "total_users": total,
            "total_players": cls.safe_count(data.get("total_players")),
            "admin_users": cls.safe_count(data.get("admin_users")),
            "super_admin_users": cls.safe_count(data.get("super_admin_users")),
            "active_users": active,
            "pending_users": cls.safe_count(data.get("pending_users")),
            "suspended_users": cls.safe_count(data.get("suspended_users")),
            "blocked_users": cls.safe_count(data.get("blocked_users")),
            "verified_users": verified,
            "unverified_users": cls.safe_count(data.get("unverified_users")),
            "new_users_in_range": current_new,
            "returning_users": cls.safe_count(data.get("returning_users")),
            "daily_active_users": cls.safe_count(data.get("daily_active_users")),
            "monthly_active_users": cls.safe_count(data.get("monthly_active_users")),
            "growth_percentage": cls.calculate_growth(current_new, previous_new),
            "active_user_rate": cls.safe_percentage(active, total),
            "verification_rate": cls.safe_percentage(verified, total),
        }

    async def get_wallet(
        self,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        granularity: AnalyticsGranularity = "day",
    ) -> WalletAnalyticsResponseSchema:
        date_range = self.resolve_date_range(
            start_date=start_date, end_date=end_date, granularity=granularity
        )
        try:
            result = await self.repository.get_wallet_analytics(
                start_date=date_range.start_date,
                end_date=date_range.end_date,
                granularity=date_range.granularity,
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsRepositoryError(
                "Unable to retrieve wallet analytics.",
                code="wallet_repository_failed",
                details={"error": str(exc)},
            ) from exc
        data = self.normalize_mapping(result)
        summary_data = self.normalize_mapping(data.get("summary"))
        total_users = await self._get_total_user_count_for_arpu(date_range=date_range)
        return WalletAnalyticsResponseSchema(
            range=date_range,
            generated_at=self.utc_now(),
            summary=self.build_wallet_summary(summary_data, total_users=total_users),
            wallet_growth=self.format_chart(
                self.normalize_sequence(data.get("wallet_growth")), count=True
            ),
            balance_trend=self.format_chart(
                self.normalize_sequence(data.get("balance_trend"))
            ),
            balance_distribution=self.format_wallet_distribution(
                self.normalize_sequence(data.get("balance_distribution"))
            ),
        )

    @classmethod
    def build_wallet_summary(
        cls, source: Mapping[str, Any] | None, *, total_users: Any = 0
    ) -> dict[str, Any]:
        data = cls.normalize_mapping(source)
        total = cls.safe_count(data.get("total_wallets"))
        frozen = cls.safe_count(data.get("frozen_wallets"))
        return {
            "total_wallets": total,
            "active_wallets": cls.safe_count(data.get("active_wallets")),
            "frozen_wallets": frozen,
            "zero_balance_wallets": cls.safe_count(data.get("zero_balance_wallets")),
            "positive_balance_wallets": cls.safe_count(data.get("positive_balance_wallets")),
            "total_coins_in_circulation": cls.safe_coin_amount(data.get("total_coins_in_circulation")),
            "average_wallet_balance": cls.round_money(data.get("average_wallet_balance")),
            "minimum_wallet_balance": cls.safe_coin_amount(data.get("minimum_wallet_balance")),
            "maximum_wallet_balance": cls.safe_coin_amount(data.get("maximum_wallet_balance")),
            "frozen_wallet_rate": cls.safe_percentage(frozen, total),
            "wallet_conversion_rate": cls.safe_percentage(total, total_users),
        }

    @classmethod
    def format_wallet_distribution(
        cls, rows: Sequence[Mapping[str, Any]] | None
    ) -> list[WalletBalanceBucketSchema]:
        normalized = cls.normalize_sequence(rows)
        total = sum(cls.safe_count(row.get("wallet_count")) for row in normalized)
        return [
            WalletBalanceBucketSchema(
                label=cls.normalize_label(row.get("label"), max_length=120),
                minimum_balance=cls.safe_coin_amount(row.get("minimum_balance")),
                maximum_balance=(
                    cls.safe_coin_amount(row.get("maximum_balance"))
                    if row.get("maximum_balance") is not None else None
                ),
                wallet_count=cls.safe_count(row.get("wallet_count")),
                percentage=cls.safe_percentage(row.get("wallet_count"), total),
            )
            for row in normalized
        ]

    async def get_transactions(
        self,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        granularity: AnalyticsGranularity = "day",
    ) -> TransactionAnalyticsResponseSchema:
        date_range = self.resolve_date_range(
            start_date=start_date, end_date=end_date, granularity=granularity
        )
        try:
            result = await self.repository.get_transaction_analytics(
                start_date=date_range.start_date,
                end_date=date_range.end_date,
                granularity=date_range.granularity,
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsRepositoryError(
                "Unable to retrieve transaction analytics.",
                code="transactions_repository_failed",
                details={"error": str(exc)},
            ) from exc
        data = self.normalize_mapping(result)
        return TransactionAnalyticsResponseSchema(
            range=date_range,
            generated_at=self.utc_now(),
            summary=self.build_transaction_summary(
                self.normalize_mapping(data.get("summary"))
            ),
            transaction_trend=self.format_chart(
                self.normalize_sequence(data.get("transaction_trend")), count=True
            ),
            amount_trend=self.format_chart(
                self.normalize_sequence(data.get("amount_trend"))
            ),
            transactions_by_type=self.format_distribution(
                self.normalize_sequence(data.get("transactions_by_type")), count=True
            ),
            amount_by_type=self.format_distribution(
                self.normalize_sequence(data.get("amount_by_type")), count=False
            ),
            hourly_distribution=self.format_hourly_distribution(
                self.normalize_sequence(data.get("hourly_distribution"))
            ),
        )

    @classmethod
    def build_transaction_summary(cls, source: Mapping[str, Any] | None) -> dict[str, Any]:
        data = cls.normalize_mapping(source)
        credited = cls.safe_coin_amount(data.get("total_credited_coins"))
        debited = cls.safe_coin_amount(data.get("total_debited_coins"))
        return {
            "total_transactions": cls.safe_count(data.get("total_transactions")),
            "purchase_transactions": cls.safe_count(data.get("purchase_transactions")),
            "game_entry_transactions": cls.safe_count(data.get("game_entry_transactions")),
            "admin_credit_transactions": cls.safe_count(data.get("admin_credit_transactions")),
            "admin_debit_transactions": cls.safe_count(data.get("admin_debit_transactions")),
            "refund_transactions": cls.safe_count(data.get("refund_transactions")),
            "total_credited_coins": credited,
            "total_debited_coins": debited,
            "net_coin_flow": cls.safe_signed_coin_amount(data.get("net_coin_flow")) if data.get("net_coin_flow") is not None else cls.calculate_net_flow(credited, debited),
            "average_transaction_amount": cls.round_money(data.get("average_transaction_amount")),
            "minimum_transaction_amount": cls.safe_coin_amount(data.get("minimum_transaction_amount")),
            "maximum_transaction_amount": cls.safe_coin_amount(data.get("maximum_transaction_amount")),
        }

    @classmethod
    def format_hourly_distribution(
        cls,
        rows: Sequence[Mapping[str, Any]] | None,
    ) -> list[dict[str, Any]]:
        row_map: dict[int, Mapping[str, Any]] = {}

        for row in cls.normalize_sequence(rows):
            raw_hour = row.get("hour")

            if isinstance(raw_hour, bool):
                continue

            try:
                hour = int(raw_hour)
            except (TypeError, ValueError, OverflowError):
                continue

            if hour < 0 or hour > 23:
                continue

            row_map[hour] = row

        return [
            {
                "hour": hour,
                "label": cls.normalize_label(
                    row_map.get(hour, {}).get("label"),
                    default=f"{hour:02d}:00",
                    max_length=20,
                ),
                "transaction_count": cls.safe_count(
                    row_map.get(hour, {}).get("transaction_count")
                ),
                "total_amount": cls.safe_coin_amount(
                    row_map.get(hour, {}).get("total_amount")
                ),
            }
            for hour in range(24)
        ]

    async def get_games(
        self,
        *,
        start_date: datetime | None = None,
        end_date: datetime | None = None,
        granularity: AnalyticsGranularity = "day",
        top_limit: int = DEFAULT_GAME_LIMIT,
    ) -> GameAnalyticsResponseSchema:
        date_range = self.resolve_date_range(
            start_date=start_date, end_date=end_date, granularity=granularity
        )
        limit = self.validate_game_limit(top_limit)
        try:
            result = await self.repository.get_game_analytics(
                start_date=date_range.start_date,
                end_date=date_range.end_date,
                granularity=date_range.granularity,
                top_limit=limit,
            )
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsRepositoryError(
                "Unable to retrieve game analytics.",
                code="games_repository_failed",
                details={"error": str(exc)},
            ) from exc
        data = self.normalize_mapping(result)
        summary = self.build_game_summary(self.normalize_mapping(data.get("summary")))
        total_plays = summary["total_play_count"]
        return GameAnalyticsResponseSchema(
            range=date_range,
            generated_at=self.utc_now(),
            summary=summary,
            game_creation_trend=self.format_chart(
                self.normalize_sequence(data.get("game_creation_trend")), count=True
            ),
            games_by_status=self.format_distribution(
                self.normalize_sequence(data.get("games_by_status")), count=True
            ),
            games_by_category=self.format_distribution(
                self.normalize_sequence(data.get("games_by_category")), count=True
            ),
            games_by_provider=self.format_distribution(
                self.normalize_sequence(data.get("games_by_provider")), count=True
            ),
            plays_by_category=self.format_distribution(
                self.normalize_sequence(data.get("plays_by_category")), count=False, total=total_plays
            ),
            plays_by_provider=self.format_distribution(
                self.normalize_sequence(data.get("plays_by_provider")), count=False, total=total_plays
            ),
            top_games=self.format_game_table(
                self.normalize_sequence(data.get("top_games")), total_plays=total_plays, limit=limit
            ),
            least_played_games=self.format_game_table(
                self.normalize_sequence(data.get("least_played_games")), total_plays=total_plays, limit=limit
            ),
        )

    @classmethod
    def build_game_summary(cls, source: Mapping[str, Any] | None) -> dict[str, Any]:
        data = cls.normalize_mapping(source)
        total = cls.safe_count(data.get("total_games"))
        published = cls.safe_count(data.get("published_games"))
        featured = cls.safe_count(data.get("featured_games"))
        plays = cls.safe_count(data.get("total_play_count"))
        return {
            "total_games": total,
            "published_games": published,
            "draft_games": cls.safe_count(data.get("draft_games")),
            "maintenance_games": cls.safe_count(data.get("maintenance_games")),
            "disabled_games": cls.safe_count(data.get("disabled_games")),
            "featured_games": featured,
            "landing_page_games": cls.safe_count(data.get("landing_page_games")),
            "total_play_count": plays,
            "average_plays_per_game": cls.safe_average(plays, total),
            "published_rate": cls.safe_percentage(published, total),
            "featured_rate": cls.safe_percentage(featured, total),
        }

    async def get_section_payload(
        self,
        section: AnalyticsSection,
        *,
        date_range: AnalyticsDateRangeSchema,
    ) -> Any:
        section = self.validate_section(section)
        kwargs = {
            "start_date": date_range.start_date,
            "end_date": date_range.end_date,
        }
        if section == "overview":
            return await self.get_overview(**kwargs)
        if section == "revenue":
            return await self.get_revenue(**kwargs, granularity=date_range.granularity)
        if section == "users":
            return await self.get_users(**kwargs, granularity=date_range.granularity)
        if section == "wallet":
            return await self.get_wallet(**kwargs, granularity=date_range.granularity)
        if section == "transactions":
            return await self.get_transactions(**kwargs, granularity=date_range.granularity)
        if section == "games":
            return await self.get_games(**kwargs, granularity=date_range.granularity)
        return {
            "overview": await self.get_overview(**kwargs),
            "revenue": await self.get_revenue(**kwargs, granularity=date_range.granularity),
            "users": await self.get_users(**kwargs, granularity=date_range.granularity),
            "wallet": await self.get_wallet(**kwargs, granularity=date_range.granularity),
            "transactions": await self.get_transactions(**kwargs, granularity=date_range.granularity),
            "games": await self.get_games(**kwargs, granularity=date_range.granularity),
        }

    @classmethod
    def generate_xlsx_buffer(cls, payload: Any, *, section: str) -> BytesIO:
        try:
            from openpyxl import Workbook
        except ImportError as exc:
            raise AnalyticsDependencyError("openpyxl", export_format="xlsx") from exc
        try:
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            serialized = cls.serialize_export_payload(payload)
            sections = serialized if isinstance(serialized, Mapping) and section == "full" else {section: serialized}
            for name, value in sections.items():
                title = str(name)[:31] or "Analytics"
                sheet = workbook.create_sheet(title=title)
                rows = cls.payload_to_flat_rows(value, section=str(name))
                headers = cls.collect_export_headers(rows)
                sheet.append(headers)
                for row in rows:
                    sheet.append([cls.normalize_export_cell(row.get(header)) for header in headers])
                sheet.freeze_panes = "A2"
                sheet.auto_filter.ref = sheet.dimensions
                for column_cells in sheet.columns:
                    width = min(max(len(str(cell.value or "")) for cell in column_cells) + 2, 60)
                    sheet.column_dimensions[column_cells[0].column_letter].width = width
            if not workbook.sheetnames:
                workbook.create_sheet("Analytics")
            buffer = BytesIO()
            workbook.save(buffer)
            return cls.rewind_buffer(buffer)
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsExportError(
                "Unable to generate the XLSX analytics export.",
                code="xlsx_export_failed",
                details={"section": section, "error": str(exc)},
            ) from exc

    @classmethod
    def generate_pdf_buffer(cls, payload: Any, *, section: str) -> BytesIO:
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        except ImportError as exc:
            raise AnalyticsDependencyError("reportlab", export_format="pdf") from exc
        try:
            buffer = BytesIO()
            document = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=24, leftMargin=24, topMargin=24, bottomMargin=24)
            styles = getSampleStyleSheet()
            story = [Paragraph(f"GoldenSweep Analytics — {section.title()}", styles["Title"]), Spacer(1, 12)]
            serialized = cls.serialize_export_payload(payload)
            sections = serialized if isinstance(serialized, Mapping) and section == "full" else {section: serialized}
            first = True
            for name, value in sections.items():
                if not first:
                    story.append(PageBreak())
                first = False
                story.append(Paragraph(str(name).replace("_", " ").title(), styles["Heading1"]))
                rows = cls.payload_to_flat_rows(value, section=str(name))
                headers = cls.collect_export_headers(rows)
                table_data = [headers] + [[str(cls.normalize_export_cell(row.get(h)))[:120] for h in headers] for row in rows]
                table = Table(table_data, repeatRows=1)
                table.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D4AF37")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ]))
                story.append(table)
            document.build(story)
            return cls.rewind_buffer(buffer)
        except AnalyticsServiceError:
            raise
        except Exception as exc:
            raise AnalyticsExportError(
                "Unable to generate the PDF analytics export.",
                code="pdf_export_failed",
                details={"section": section, "error": str(exc)},
            ) from exc

    async def export_analytics(
        self,
        request: AnalyticsExportRequestSchema,
    ) -> tuple[BytesIO, AnalyticsExportResponseSchema]:
        export_format = self.validate_export_format(request.format)
        section = self.validate_section(request.section)
        date_range = self.resolve_export_date_range(request)
        generated_at = self.utc_now()
        payload = await self.get_section_payload(section, date_range=date_range)
        if export_format == "csv":
            buffer = self.generate_csv_buffer(payload, section=section)
        elif export_format == "xlsx":
            buffer = self.generate_xlsx_buffer(payload, section=section)
        else:
            buffer = self.generate_pdf_buffer(payload, section=section)
        buffer = self.validate_export_buffer(buffer, export_format=export_format)
        filename = self.build_export_filename(
            section=section,
            export_format=export_format,
            generated_at=generated_at,
        )
        metadata = self.build_export_response(
            export_format=export_format,
            section=section,
            filename=filename,
            generated_at=generated_at,
        )
        return buffer, metadata

    async def generate_export(
        self,
        request: AnalyticsExportRequestSchema,
    ) -> tuple[BytesIO, AnalyticsExportResponseSchema]:
        return await self.export_analytics(request)


analytics_service = AnalyticsService()


__all__ = [
    "AnalyticsSection",
    "GrowthDirection",
    "AnalyticsServiceError",
    "AnalyticsValidationError",
    "AnalyticsRepositoryError",
    "AnalyticsExportError",
    "AnalyticsDependencyError",
    "AnalyticsService",
    "analytics_service",
]